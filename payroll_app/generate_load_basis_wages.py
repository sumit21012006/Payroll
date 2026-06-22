import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random

# Seed for reproducible random numbers
random.seed(42)

# Load Master KFIL
master_wb = openpyxl.load_workbook("d:/DEEPTI_ANTIGRAVITY/Real Data/Master KFIL.xlsx")
master_sheet = master_wb.active

# Parse employees in load basis departments
load_basis_depts = {'HE', 'FINAL', 'REWORK', 'PAINTER', 'AVG', 'YANMAR LINE'}
employees = []
for r in range(2, master_sheet.max_row + 1):
    dept = master_sheet.cell(row=r, column=7).value
    if dept:
        dept_clean = dept.strip().upper()
        if dept_clean in load_basis_depts:
            emp = {
                'ticket': master_sheet.cell(row=r, column=3).value,
                'uan': master_sheet.cell(row=r, column=4).value,
                'esic': master_sheet.cell(row=r, column=5).value,
                'name': master_sheet.cell(row=r, column=6).value,
                'dept': dept_clean,
            }
            employees.append(emp)

print(f"Loaded {len(employees)} load-basis employees.")

# Load Wages Register Template
wages_wb = openpyxl.load_workbook("d:/DEEPTI_ANTIGRAVITY/Real Data/Demo File For Sallay Wages.xlsx")
wages_sheet = wages_wb.active
wages_sheet.title = "(PF) WAGESH MAY-26 Load-01"

# Rename the header title cell if any
title_cell = wages_sheet['A5']
if title_cell.value:
    title_cell.value = str(title_cell.value).replace("APR-26", "MAY-26")

# Let's save a template row styles to copy later
template_cells = []
for col in range(1, 28):
    template_cells.append(wages_sheet.cell(row=12, column=col))

# Save total row styles from row 122 of original workbook
total_template_cells = []
for col in range(1, 28):
    total_template_cells.append(wages_sheet.cell(row=122, column=col))

# Clear all data rows from row 12 onwards (up to 130)
for r in range(12, 130):
    for c in range(1, 28):
        wages_sheet.cell(row=r, column=c).value = None

# Let's write employees data
row_idx = 12
sr_no = 1

def copy_style(src_cell, dest_cell):
    if src_cell.font:
        dest_cell.font = Font(name=src_cell.font.name, size=src_cell.font.size, bold=src_cell.font.bold, italic=src_cell.font.italic, color=src_cell.font.color)
    if src_cell.fill:
        dest_cell.fill = PatternFill(fill_type=src_cell.fill.fill_type, start_color=src_cell.fill.start_color, end_color=src_cell.fill.end_color)
    if src_cell.border:
        dest_cell.border = Border(left=src_cell.border.left, right=src_cell.border.right, top=src_cell.border.top, bottom=src_cell.border.bottom)
    if src_cell.alignment:
        dest_cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical, wrap_text=src_cell.alignment.wrap_text)
    dest_cell.number_format = src_cell.number_format

# Database details for the second sheet
db_records = []

for emp in employees:
    dept = emp['dept']
    # Simulate work
    days_worked = random.randint(18, 26)
    ot_days = random.randint(2, 8) if random.random() > 0.4 else 0
    
    # Rates and quantity simulation
    rate = 0.0
    qty = 0.0
    unit = 'Tons'
    if dept == 'HE':
        rate = 320.0
        qty = round(random.uniform(50.0, 90.0), 3)
        unit = 'Tons'
    elif dept == 'FINAL':
        rate = 220.0
        qty = round(random.uniform(70.0, 120.0), 3)
        unit = 'Tons'
    elif dept == 'REWORK':
        rate = 4.90
        qty = random.randint(3000, 5000)
        unit = 'Pieces'
    elif dept == 'PAINTER':
        rate = 6.00
        qty = random.randint(2500, 4500)
        unit = 'Pieces'
    elif dept == 'AVG':
        rate = 5.00
        qty = random.randint(3000, 5000)
        unit = 'Pieces'
    elif dept == 'YANMAR LINE':
        rate = 28.00
        qty = random.randint(600, 1000)
        unit = 'Pieces'
        
    gross_earnings = qty * rate
    effective_rate = round(gross_earnings / days_worked, 2)
    
    # Store for db sheet
    db_records.append({
        'ticket': emp['ticket'],
        'name': emp['name'],
        'dept': dept,
        'unit': unit,
        'rate': rate,
        'qty': qty,
        'earnings': gross_earnings,
        'days': days_worked
    })

    # Write to main wages register sheet
    row_vals = [
        sr_no,
        emp['uan'],
        emp['esic'],
        emp['name'],
        effective_rate,
        days_worked,
        ot_days if ot_days > 0 else None,
        f"=F{row_idx}+G{row_idx}" if ot_days > 0 else f"=F{row_idx}",
        f"=ROUND(F{row_idx}*E{row_idx},0)",
        f"=ROUND(E{row_idx}*G{row_idx},0)" if ot_days > 0 else 0,
        f"=I{row_idx}+J{row_idx}",
        f"=ROUND($L$4*F{row_idx},0)",
        f"=ROUND($L{row_idx}*5%,0)",
        f"=K{row_idx}-L{row_idx}-M{row_idx}",
        f"=N{row_idx}+M{row_idx}+L{row_idx}",
        f"=ROUND(L{row_idx}*12%,0)",
        f"=IF(O{row_idx}<=7500,0,IF(O{row_idx}<=10000,175,200))",
        f"=ROUND(O{row_idx}*0.75%,0)",
        None, # Adv Remaining
        None, # Adv Jamadar
        None, # Canteen
        None, # Cupan
        None, # Account Adv
        f"=SUM(P{row_idx}:W{row_idx})",
        f"=O{row_idx}-X{row_idx}",
        500, # canteen other
        f"=Y{row_idx}-Z{row_idx}"
    ]

    for col_idx in range(1, 28):
        cell = wages_sheet.cell(row=row_idx, column=col_idx)
        cell.value = row_vals[col_idx - 1]
        copy_style(template_cells[col_idx - 1], cell)

    sr_no += 1
    row_idx += 1

# Write total Row
total_row_idx = row_idx
wages_sheet.cell(row=total_row_idx, column=4).value = "TOTAL"

# Apply totals for other columns
for col_idx in [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 24, 25, 26, 27]:
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    wages_sheet.cell(row=total_row_idx, column=col_idx).value = f"=SUM({col_letter}12:{col_letter}{total_row_idx-1})"

# Style total Row
for col_idx in range(1, 28):
    cell = wages_sheet.cell(row=total_row_idx, column=col_idx)
    copy_style(total_template_cells[col_idx - 1], cell)

# Create the database sheet
db_sheet = wages_wb.create_sheet(title="Load Basis Work Database")
db_sheet.views.sheetView[0].showGridLines = True

# Headers for database sheet
db_headers = [
    'Sr No', 'Ticket No', 'UAN NO', 'Employee Name', 'Department', 
    'Unit Type', 'Rate per Unit (₹)', 'Total Qty Processed', 'Calculated Load Earnings (₹)', 'Days Worked'
]

# Style for db sheet header
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(fill_type='solid', start_color='1F2937', end_color='1F2937')
header_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

for col_idx, h in enumerate(db_headers, 1):
    cell = db_sheet.cell(row=1, column=col_idx)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write db sheet records
for idx, rec in enumerate(db_records, 1):
    row_num = idx + 1
    uan = next((e['uan'] for e in employees if e['ticket'] == rec['ticket']), '')
    
    row_vals = [
        idx,
        rec['ticket'],
        uan,
        rec['name'],
        rec['dept'],
        rec['unit'],
        rec['rate'],
        rec['qty'],
        f"=G{row_num}*H{row_num}",
        rec['days']
    ]
    
    for col_idx, val in enumerate(row_vals, 1):
        cell = db_sheet.cell(row=row_num, column=col_idx)
        cell.value = val
        cell.border = thin_border
        if col_idx in [1, 2, 3, 5, 6, 10]:
            cell.alignment = Alignment(horizontal='center')
        elif col_idx in [7, 8, 9]:
            cell.alignment = Alignment(horizontal='right')
            if col_idx in [7, 9]:
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '#,##0.000' if rec['unit'] == 'Tons' else '#,##0'

# Auto-fit column widths for database sheet
for col in db_sheet.columns:
    max_len = 0
    for cell in col:
        val_str = str(cell.value or '')
        # Simple length check
        if len(val_str) > max_len:
            max_len = len(val_str)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    db_sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save the workbook
output_path = "d:/DEEPTI_ANTIGRAVITY/Real Data/Demo File For Load Basis Wages.xlsx"
wages_wb.save(output_path)
print(f"Generated Load-Basis Excel sheet at: {output_path}")
